# ---------------------------------------------------------
# IMPORTS
# ---------------------------------------------------------

# FastAPI core
from fastapi import APIRouter, Depends, HTTPException

# Norm business logic methods
from manage.services.norms import methods as normMethods

# Response schemas
from manage.services.norms.schema import FacilityResult, NodeRollup

# Database session factory
from manage.database import SessionLocal, engine

# Organization unit helpers
from manage.services.organization_units import methods as orgUnitMethods

# Async DB session
from sqlalchemy.ext.asyncio import AsyncSession

# Raw SQL (not directly used here but may be required in future extensions)
from sqlalchemy import text

# Streaming response for file export
from fastapi.responses import StreamingResponse

# In-memory stream handling
from io import BytesIO

# Date formatting
from datetime import datetime

# Excel generation
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment


# Create router
apiRouter = APIRouter()


# ---------------------------------------------------------
# Dependency: Get Async Database Session
# ---------------------------------------------------------
async def get_session() -> AsyncSession:
    """
    Provides an async DB session.
    Ensures proper open/close lifecycle handling.
    """
    async with SessionLocal() as session:
        yield session


# =========================================================
# FACILITY-LEVEL NORMS
# =========================================================

@apiRouter.get("/norms/facilities", response_model=list[FacilityResult])
async def check_all_facilities(db: AsyncSession = Depends(get_session)):
    """
    Compute staffing norms compliance for ALL facilities.

    For each facility:
    - Get its organization unit type
    - Load required staffing norms
    - Load actual staff counts
    - Compute missing / excess / compliance ratio
    """

    facilities = await normMethods.db_list_facilities(db)

    out = []

    for f in facilities:
        # Get facility type (HD, CS_RURAL, etc.)
        organization_unit_type_id = await normMethods.db_get_organization_unit_type_id(f["id"], db)

        # Required staffing norms
        required = await normMethods.db_get_norms(organization_unit_type_id, db)

        # Actual staff deployed
        actual = await normMethods.db_get_actual_staff(f["id"], db)

        if actual is None:
            actual = {}

        # Compute missing, excess, compliance
        calc = normMethods.compute_facility(required, actual)

        out.append(FacilityResult(
            org_unit_id=f["id"],
            organization_unit_type_id=organization_unit_type_id,
            required=required,
            actual=actual,
            missing=calc["missing"],
            excess=calc["excess"],
            compliance_ratio=calc["compliance_ratio"]
        ))

    return out


# =========================================================
# NODE ROLLUP (DISTRICT / PROVINCE AGGREGATION)
# =========================================================

@apiRouter.get("/norms/rollup/{node_id}", response_model=NodeRollup)
async def rollup_node(node_id: str, db: AsyncSession = Depends(get_session)):
    """
    Compute aggregated norms compliance for a parent node
    (e.g., District or Province).

    Logic:
    - Retrieve all descendant facilities
    - Sum required staffing
    - Sum actual staffing
    - Compute global compliance ratio
    """

    facilities = await normMethods.db_list_descendant_facilities(node_id, db)

    if not facilities:
        raise HTTPException(404, "No descendant facilities found")

    required_list = []
    actual_list = []

    for facilityId in facilities:
        organization_unit_type_id = await normMethods.db_get_organization_unit_type_id(facilityId, db)
        required = await normMethods.db_get_norms(organization_unit_type_id, db)
        actual = await normMethods.db_get_actual_staff(facilityId, db)

        required_list.append(required)
        actual_list.append(actual)

    # Aggregate all facilities
    required_sum = normMethods.sum_dicts(required_list)
    actual_sum = normMethods.sum_dicts(actual_list)

    # Compute aggregated stats
    calc = normMethods.compute_facility(required_sum, actual_sum)

    required_total = sum(required_sum.values())
    actual_total = sum(actual_sum.values())
    missing_total = sum(calc["missing"].values())

    return NodeRollup(
        org_unit_id=node_id,
        level="aggregated",
        required_total=required_total,
        actual_total=actual_total,
        missing_total=missing_total,
        compliance_ratio=calc["compliance_ratio"]
    )


# =========================================================
# BUILD FULL NORM TREE REPORT
# =========================================================

@apiRouter.get("/norms/{node_id}/tree")
async def check_norms_from_node(node_id: str, db: AsyncSession = Depends(get_session)):
    """
    Returns a full hierarchical norms report (tree structure)
    starting from a given node.
    """
    return await normMethods.build_norm_tree_report(node_id, db)


# =========================================================
# EXCEL EXPORT HELPERS
# =========================================================

def _autofit(ws):
    """
    Auto-adjust Excel column widths based on content length.
    """
    for col in range(1, ws.max_column + 1):
        max_len = 0
        letter = get_column_letter(col)

        for row in range(1, ws.max_row + 1):
            value = ws.cell(row=row, column=col).value
            if value is None:
                continue
            max_len = max(max_len, len(str(value)))

        ws.column_dimensions[letter].width = min(max_len + 2, 60)


def _all_keys(stat: dict) -> list[str]:
    """
    Extract all classification IDs from:
    - required
    - actual
    - missing_total
    - excess
    """
    keys = set()

    if not stat:
        return []

    for k in ("required", "actual", "missing_total", "excess"):
        obj = stat.get(k) or {}
        keys.update(obj.keys())

    return sorted(keys)


def _write_orgunit_table(ws, start_row: int, org_unit: dict, classification_map: dict | None = None) -> int:
    """
    Write norms table for one organization unit into Excel.

    Layout:
    NORMES - <OrgUnitName>
    CATEGORIE | REQUIS | ACTIF | CARENCE | PLETHORE
    """

    stat = org_unit.get("stat") or {}

    # Title row
    title = f"NORMES - {org_unit.get('name')}"
    ws.cell(row=start_row, column=1, value=title).font = Font(bold=True, size=12)
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=5)

    # Header row
    header_row = start_row + 1
    headers = ["CATEGORIE", "REQUIS", "ACTIF", "CARENCE", "PLETHORE"]

    for i, h in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=i, value=h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    row = header_row + 1

    # Write data rows
    for cid in _all_keys(stat):
        name = cid

        # Replace classification ID with readable name if available
        if classification_map and cid in classification_map:
            name = classification_map[cid].get("name") or cid

        ws.cell(row=row, column=1, value=name)
        ws.cell(row=row, column=2, value=int((stat.get("required") or {}).get(cid, 0) or 0))
        ws.cell(row=row, column=3, value=int((stat.get("actual") or {}).get(cid, 0) or 0))
        ws.cell(row=row, column=4, value=int((stat.get("missing_total") or {}).get(cid, 0) or 0))
        ws.cell(row=row, column=5, value=int((stat.get("excess") or {}).get(cid, 0) or 0))

        row += 1

    return row + 2


# =========================================================
# EXPORT TREE REPORT TO EXCEL
# =========================================================

@apiRouter.get("/norms/{node_id}/tree/export")
async def export_norms_excel(node_id: str, db: AsyncSession = Depends(get_session)):
    """
    Export hierarchical norms report to Excel (.xlsx).

    Includes:
    - Node metadata (name, id, timestamp)
    - Norm tables for each org unit in tree
    """

    report = await normMethods.build_norm_tree_report(node_id, db)
    tree = report.get("tree") or []

    wb = Workbook()
    ws = wb.active
    ws.title = "Normes"

    # Metadata header
    node = report.get("node") or {}

    ws["A1"] = "Node"
    ws["B1"] = node.get("name")

    ws["A2"] = "Node ID"
    ws["B2"] = node.get("id")

    ws["A3"] = "Generated at"
    ws["B3"] = datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S UTC")

    ws["A1"].font = ws["A2"].font = ws["A3"].font = Font(bold=True)

    r = 5

    # Load classification labels for display
    classificationMap = await normMethods.db_get_classification_map(db)

    for org_unit in tree:
        r = _write_orgunit_table(ws, r, org_unit, classificationMap)

    _autofit(ws)

    # Stream file
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    filename = f"norms_{node_id}.xlsx"

    return StreamingResponse(
        bio,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )